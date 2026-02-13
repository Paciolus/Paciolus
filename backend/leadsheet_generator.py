"""
Paciolus Lead Sheet Generator
Sprint 25: Automated Lead Sheet / Working Paper Generator

Generates standardized account-level working papers (Lead Sheets) using 
Flux and Recon data. 

Zero-Storage Compliance:
- Consumes FluxResult and ReconResult (in-memory)
- Generates Excel file in BytesIO buffer
- No disk I/O

Features:
- "Lead Sheet Index": Master summary of all accounts with tickmarks/risk
- "High Risk Accounts": Detailed tabs for accounts needing attention
- Rollforward logic: Beg Bal + Activity = End Bal
"""

import io
from datetime import datetime, UTC
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter

from security_utils import log_secure_operation
from excel_generator import ExcelColors, create_header_style, create_currency_style
from shared.helpers import sanitize_csv_value
from flux_engine import FluxResult, FluxItem
from recon_engine import ReconResult, ReconScore, RiskBand

class LeadSheetGenerator:
    """
    Generates Excel Lead Sheets from diagnostic results.
    """

    def __init__(self, flux_result: FluxResult, recon_result: ReconResult, filename: str = "LeadSheets"):
        self.flux_result = flux_result
        self.recon_result = recon_result
        self.filename = filename
        self.wb = Workbook()
        self.buffer = io.BytesIO()
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Initialize styles
        self._register_styles()
        
        # Map recon scores for easy lookup
        self.recon_map = {s.account_name: s for s in recon_result.scores}

    def _register_styles(self):
        """Register specific styles for lead sheets."""
        # Reuse or create styles
        try:
            self.wb.add_named_style(create_header_style())
            self.wb.add_named_style(create_currency_style())
        except ValueError:
            pass # Styles might already exist if reusing global instance logic
            
        # Lead Sheet Title
        title_style = NamedStyle(name="ls_title")
        title_style.font = Font(bold=True, size=14, color=ExcelColors.OBSIDIAN)
        self.wb.add_named_style(title_style)
        
        # Risk High
        risk_high = NamedStyle(name="risk_high")
        risk_high.font = Font(bold=True, color="FFFFFF")
        risk_high.fill = PatternFill("solid", fgColor=ExcelColors.CLAY)
        risk_high.alignment = Alignment(horizontal="center")
        self.wb.add_named_style(risk_high)

        # Risk Medium
        risk_med = NamedStyle(name="risk_med")
        risk_med.font = Font(bold=True, color=ExcelColors.OBSIDIAN)
        risk_med.fill = PatternFill("solid", fgColor=ExcelColors.OATMEAL_400)
        risk_med.alignment = Alignment(horizontal="center")
        self.wb.add_named_style(risk_med)

    def generate(self) -> bytes:
        """Generate the workbook and return bytes."""
        log_secure_operation("leadsheet_gen", f"Generating lead sheets for {len(self.flux_result.items)} accounts")
        
        self._build_index_tab()
        self._build_detail_tabs()
        
        self.wb.save(self.buffer)
        data = self.buffer.getvalue()
        self.buffer.close()
        return data

    def _build_index_tab(self):
        """Create the Master Level Lead Sheet (Index)."""
        ws = self.wb.create_sheet("Lead Sheet Index", 0)
        
        # Header Info
        ws['A1'] = "Lead Sheet Summary / Index"
        ws['A1'].style = "ls_title"
        ws['A2'] = f"Client File: {self.filename}"
        ws['A3'] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d')}"
        
        # Table Headers
        headers = [
            "Account", "Type", 
            "Prior Year (PY)", "Current Year (CY)", "Variance ($)", "Var (%)", 
            "Recon Risk", "Score", "Notes / Tickmarks"
        ]
        
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=text)
            cell.style = "header_style"
            
        # Data Rows
        row_idx = 6
        for item in self.flux_result.items:
            recon = self.recon_map.get(item.account_name)
            
            # Account Name
            ws.cell(row=row_idx, column=1, value=sanitize_csv_value(item.account_name))
            ws.cell(row=row_idx, column=2, value=sanitize_csv_value(item.account_type))
            
            # Financials
            c_py = ws.cell(row=row_idx, column=3, value=item.prior_balance)
            c_py.style = "currency_style"
            
            c_cy = ws.cell(row=row_idx, column=4, value=item.current_balance)
            c_cy.style = "currency_style"
            
            c_var = ws.cell(row=row_idx, column=5, value=item.delta_amount)
            c_var.style = "currency_style"
            
            # Variance %
            pct_val = item.delta_percent / 100.0 if item.prior_balance != 0 else 0
            c_pct = ws.cell(row=row_idx, column=6, value=pct_val)
            c_pct.number_format = '0.0%'
            
            # Risk Info
            if recon:
                risk_cell = ws.cell(row=row_idx, column=7, value=recon.risk_band.value.upper())
                ws.cell(row=row_idx, column=8, value=recon.risk_score)
                
                if recon.risk_band == RiskBand.HIGH:
                    risk_cell.style = "risk_high"
                elif recon.risk_band == RiskBand.MEDIUM:
                    risk_cell.style = "risk_med"
            else:
                 ws.cell(row=row_idx, column=7, value="-")
                 ws.cell(row=row_idx, column=8, value="-")
            
            # Annotations (from Flux/Recon factors)
            notes = []
            if item.is_new_account: notes.append("[New]")
            if item.has_sign_flip: notes.append("[SignFlip]")
            if recon and recon.factors:
                notes.extend(recon.factors)
            
            ws.cell(row=row_idx, column=9, value=", ".join(notes))
            
            row_idx += 1
            
        # Columns sizing
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['I'].width = 50

    def _build_detail_tabs(self):
        """Create individual tabs for High Risk accounts."""
        high_risk_items = []
        for item in self.flux_result.items:
            recon = self.recon_map.get(item.account_name)
            if recon and recon.risk_band == RiskBand.HIGH:
                high_risk_items.append(item)
                
        # Limit to top 20 to avoid generating too many tabs
        high_risk_items = high_risk_items[:20]
        
        for item in high_risk_items:
            # Excel sheet names max 31 chars
            safe_name = "".join(c for c in item.account_name if c.isalnum() or c in (' ','-','_'))
            sheet_name = safe_name[:30] 
            
            # Handle duplicate sheet names
            if sheet_name in self.wb.sheetnames:
                continue
                
            ws = self.wb.create_sheet(sheet_name)
            recon = self.recon_map.get(item.account_name)
            
            # Header
            ws['A1'] = f"Lead Sheet: {sanitize_csv_value(item.account_name)}"
            ws['A1'].style = "ls_title"
            ws['A2'] = f"Risk Assessment: {recon.risk_band.value.upper() if recon else 'N/A'}"
            if recon and recon.risk_band == RiskBand.HIGH:
                ws['A2'].style = "risk_high"
            
            # Rollforward Table
            ws['A5'] = "Beginning Balance (PY)"
            ws['B5'] = item.prior_balance
            ws['B5'].style = "currency_style"
            
            ws['A6'] = "Activity / Variance"
            ws['B6'] = item.delta_amount
            ws['B6'].style = "currency_style"
            
            ws['A7'] = "Ending Balance (CY)"
            ws['B7'] = item.current_balance
            ws['B7'].style = "currency_style"
            ws['B7'].font = Font(bold=True)
            ws['B7'].border = Border(top=Side(style='thin'), bottom=Side(style='double'))
            
            # Risk Factors
            ws['A10'] = "Detected Risk Factors:"
            ws['A10'].font = Font(bold=True)
            
            row = 11
            if recon and recon.factors:
                for factor in recon.factors:
                 ws.cell(row=row, column=1, value=f"• {factor}")
                 row += 1
            if item.risk_reasons:
                for reason in item.risk_reasons:
                    ws.cell(row=row, column=1, value=f"• Flux: {reason}")
                    row += 1
                    
            # Audit Procedures / Notes
            row += 2
            ws['A'+str(row)] = "Auditor Notes / Procedures Performed:"
            ws['A'+str(row)].font = Font(bold=True)
            
            # Make columns wide
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20

def generate_leadsheets(flux_result: FluxResult, recon_result: ReconResult, filename: str) -> bytes:
    """Public wrapper."""
    generator = LeadSheetGenerator(flux_result, recon_result, filename)
    return generator.generate()
