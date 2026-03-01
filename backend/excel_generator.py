"""Multi-tab Excel workpaper generator using Oat & Obsidian theme."""

import io
from datetime import UTC, datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from security_utils import log_secure_operation
from shared.helpers import sanitize_csv_value


class ExcelColors:
    """Oat & Obsidian theme colors for Excel workpapers."""

    # Core palette (RGB format for openpyxl)
    OBSIDIAN = "212121"
    OBSIDIAN_700 = "303030"
    OBSIDIAN_600 = "424242"
    OBSIDIAN_500 = "616161"

    OATMEAL = "EBE9E4"
    OATMEAL_300 = "DDD9D1"
    OATMEAL_400 = "C9C3B8"
    OATMEAL_500 = "B5AD9F"

    CLAY = "BC4749"
    CLAY_400 = "D16C6E"
    CLAY_600 = "A33D3F"

    SAGE = "4A7C59"
    SAGE_400 = "6FA882"
    SAGE_600 = "3D6649"

    WHITE = "FFFFFF"
    LIGHT_GRAY = "F5F4F2"


def create_header_style() -> NamedStyle:
    """Create header row style with Obsidian background."""
    style = NamedStyle(name="header_style")
    style.font = Font(bold=True, color=ExcelColors.OATMEAL, size=11)
    style.fill = PatternFill("solid", fgColor=ExcelColors.OBSIDIAN)
    style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    style.border = Border(bottom=Side(style="thin", color=ExcelColors.OATMEAL_400))
    return style


def create_title_style() -> NamedStyle:
    """Create title style for section headers."""
    style = NamedStyle(name="title_style")
    style.font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=14)
    style.alignment = Alignment(horizontal="left", vertical="center")
    return style


def create_subtitle_style() -> NamedStyle:
    """Create subtitle style."""
    style = NamedStyle(name="subtitle_style")
    style.font = Font(color=ExcelColors.OBSIDIAN_500, size=10, italic=True)
    style.alignment = Alignment(horizontal="left", vertical="center")
    return style


def create_currency_style() -> NamedStyle:
    """Create currency format style."""
    style = NamedStyle(name="currency_style")
    style.number_format = '"$"#,##0.00'
    style.font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
    style.alignment = Alignment(horizontal="right")
    return style


def create_percent_style() -> NamedStyle:
    """Create percentage format style."""
    style = NamedStyle(name="percent_style")
    style.number_format = "0.00%"
    style.font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
    style.alignment = Alignment(horizontal="right")
    return style


def create_material_style() -> NamedStyle:
    """Create style for material anomalies (Clay accent)."""
    style = NamedStyle(name="material_style")
    style.font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
    style.fill = PatternFill("solid", fgColor=ExcelColors.LIGHT_GRAY)
    style.border = Border(left=Side(style="thick", color=ExcelColors.CLAY))
    return style


def create_immaterial_style() -> NamedStyle:
    """Create style for immaterial anomalies."""
    style = NamedStyle(name="immaterial_style")
    style.font = Font(color=ExcelColors.OBSIDIAN_500, size=10)
    style.fill = PatternFill("solid", fgColor=ExcelColors.LIGHT_GRAY)
    return style


def create_balanced_style() -> NamedStyle:
    """Create style for balanced status (Sage)."""
    style = NamedStyle(name="balanced_style")
    style.font = Font(bold=True, color=ExcelColors.SAGE, size=12)
    style.alignment = Alignment(horizontal="center")
    return style


def create_unbalanced_style() -> NamedStyle:
    """Create style for unbalanced status (Clay)."""
    style = NamedStyle(name="unbalanced_style")
    style.font = Font(bold=True, color=ExcelColors.CLAY, size=12)
    style.alignment = Alignment(horizontal="center")
    return style


class PaciolusWorkpaperGenerator:
    """Generates multi-tab Excel workpapers using BytesIO buffer."""

    def __init__(
        self,
        audit_result: dict[str, Any],
        filename: str = "workpaper",
        prepared_by: Optional[str] = None,
        reviewed_by: Optional[str] = None,
        workpaper_date: Optional[str] = None,
        include_signoff: bool = False,
    ):
        self.audit_result = audit_result
        self.filename = filename
        self.wb = Workbook()
        self.buffer = io.BytesIO()
        # Sprint 53: Workpaper fields (deprecated Sprint 7 — gated by include_signoff)
        self.prepared_by = prepared_by
        self.reviewed_by = reviewed_by
        self.workpaper_date = workpaper_date or datetime.now().strftime("%Y-%m-%d")
        self.include_signoff = include_signoff

        # Remove default sheet (will create named sheets)
        self.wb.remove(self.wb.active)

        # Register styles
        self._register_styles()

        log_secure_operation("excel_generator_init", f"Initializing Excel workpaper for: {filename}")

    def _register_styles(self) -> None:
        """Register all named styles with the workbook."""
        styles = [
            create_header_style(),
            create_title_style(),
            create_subtitle_style(),
            create_currency_style(),
            create_percent_style(),
            create_material_style(),
            create_immaterial_style(),
            create_balanced_style(),
            create_unbalanced_style(),
        ]

        for style in styles:
            try:
                self.wb.add_named_style(style)
            except ValueError:
                # Style already exists
                pass

    def generate(self) -> bytes:
        """Generate the Excel workpaper. Returns bytes that can be streamed directly."""
        log_secure_operation("excel_generate_start", "Starting Excel generation")

        # Build all tabs
        self._build_summary_tab()
        self._build_standardized_tb_tab()
        self._build_anomalies_tab()
        self._build_ratios_tab()

        # Save to buffer
        self.wb.save(self.buffer)
        excel_bytes = self.buffer.getvalue()
        self.buffer.close()

        log_secure_operation("excel_generate_complete", f"Excel generated: {len(excel_bytes)} bytes")

        return excel_bytes

    def _build_summary_tab(self) -> None:
        """Build the Summary tab with executive overview."""
        ws = self.wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Paciolus Diagnostic Summary"
        ws["A1"].style = "title_style"
        ws.merge_cells("A1:D1")

        # Subtitle
        ws["A2"] = f"Analysis Intelligence Report for {self.filename}"
        ws["A2"].style = "subtitle_style"
        ws.merge_cells("A2:D2")

        # Timestamp
        ws["A3"] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        ws["A3"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        # Balance Status
        is_balanced = self.audit_result.get("balanced", False)
        ws["A5"] = "Balance Status:"
        ws["B5"] = "BALANCED" if is_balanced else "OUT OF BALANCE"
        ws["B5"].style = "balanced_style" if is_balanced else "unbalanced_style"

        # Key Metrics Section
        ws["A7"] = "Key Metrics"
        ws["A7"].style = "title_style"
        ws.merge_cells("A7:B7")

        metrics = [
            ("Total Debits", self.audit_result.get("total_debits", 0), True),
            ("Total Credits", self.audit_result.get("total_credits", 0), True),
            ("Difference", self.audit_result.get("difference", 0), True),
            ("Rows Analyzed", self.audit_result.get("row_count", 0), False),
            ("Materiality Threshold", self.audit_result.get("materiality_threshold", 0), True),
            ("Material Anomalies", self.audit_result.get("material_count", 0), False),
            ("Immaterial Anomalies", self.audit_result.get("immaterial_count", 0), False),
        ]

        for i, (label, value, is_currency) in enumerate(metrics, start=9):
            ws[f"A{i}"] = label
            ws[f"A{i}"].font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            ws[f"B{i}"] = value
            if is_currency:
                ws[f"B{i}"].style = "currency_style"
            else:
                ws[f"B{i}"].alignment = Alignment(horizontal="right")

        # Consolidated info if applicable
        if self.audit_result.get("is_consolidated"):
            row = len(metrics) + 10
            ws[f"A{row}"] = "Sheets Consolidated"
            ws[f"B{row}"] = self.audit_result.get("sheet_count", 0)

        # Sprint 53: Workpaper signoff section (deprecated Sprint 7 — gated by include_signoff)
        signoff_start = 17
        if self.include_signoff and (self.prepared_by or self.reviewed_by):
            ws[f"A{signoff_start}"] = "Workpaper Sign-Off"
            ws[f"A{signoff_start}"].style = "title_style"
            ws.merge_cells(f"A{signoff_start}:B{signoff_start}")

            row = signoff_start + 1
            if self.prepared_by:
                ws[f"A{row}"] = "Prepared By:"
                ws[f"A{row}"].font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                ws[f"B{row}"] = self.prepared_by
                ws[f"C{row}"] = self.workpaper_date
                row += 1

            if self.reviewed_by:
                ws[f"A{row}"] = "Reviewed By:"
                ws[f"A{row}"].font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                ws[f"B{row}"] = self.reviewed_by
                ws[f"C{row}"] = self.workpaper_date
                row += 1

            disclaimer_start = row + 2
        else:
            disclaimer_start = 20

        # Legal disclaimer
        ws[f"A{disclaimer_start}"] = (
            "DISCLAIMER: This output is generated by an automated analytical system and supports"
        )
        ws[f"A{disclaimer_start + 1}"] = (
            "internal evaluation and professional judgment. It does not constitute an audit, review,"
        )
        ws[f"A{disclaimer_start + 2}"] = "or attestation engagement and provides no assurance."
        for row in [disclaimer_start, disclaimer_start + 1, disclaimer_start + 2]:
            ws[f"A{row}"].font = Font(color=ExcelColors.OBSIDIAN_500, size=8, italic=True)

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _build_standardized_tb_tab(self) -> None:
        """Build the Standardized TB tab with formatted trial balance."""
        ws = self.wb.create_sheet("Standardized TB", 1)

        # Title
        ws["A1"] = "Standardized Trial Balance"
        ws["A1"].style = "title_style"
        ws.merge_cells("A1:F1")

        # Headers
        headers = ["Account", "Category", "Debit", "Credit", "Net", "Classification"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = "header_style"

        # Data from abnormal balances (represents accounts with issues)
        abnormal_balances = self.audit_result.get("abnormal_balances", [])

        row = 4
        for ab in abnormal_balances:
            ws.cell(row=row, column=1, value=sanitize_csv_value(ab.get("account", "Unknown")))
            ws.cell(row=row, column=2, value=sanitize_csv_value(ab.get("type", "Unknown")))

            debit = ab.get("debit", 0) or 0
            credit = ab.get("credit", 0) or 0
            amount = ab.get("amount", 0)

            ws.cell(row=row, column=3, value=debit)
            ws.cell(row=row, column=3).style = "currency_style"

            ws.cell(row=row, column=4, value=credit)
            ws.cell(row=row, column=4).style = "currency_style"

            ws.cell(row=row, column=5, value=amount)
            ws.cell(row=row, column=5).style = "currency_style"

            materiality = ab.get("materiality", "unknown")
            ws.cell(row=row, column=6, value=materiality.upper())

            # Apply row styling based on materiality
            if materiality == "material":
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=ExcelColors.LIGHT_GRAY)
                ws.cell(row=row, column=1).border = Border(left=Side(style="thick", color=ExcelColors.CLAY))

            row += 1

        # Totals row
        row += 1
        ws.cell(row=row, column=1, value="TOTALS")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=self.audit_result.get("total_debits", 0))
        ws.cell(row=row, column=3).style = "currency_style"
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=self.audit_result.get("total_credits", 0))
        ws.cell(row=row, column=4).style = "currency_style"
        ws.cell(row=row, column=4).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

    def _build_anomalies_tab(self) -> None:
        """Build the Flagged Anomalies tab with detailed anomaly information."""
        ws = self.wb.create_sheet("Flagged Anomalies", 2)

        # Title
        ws["A1"] = "Flagged Anomalies"
        ws["A1"].style = "title_style"
        ws.merge_cells("A1:G1")  # Sprint 53: Extended for Ref column

        # Subtitle
        ws["A2"] = f"Materiality Threshold: ${self.audit_result.get('materiality_threshold', 0):,.2f}"
        ws["A2"].style = "subtitle_style"

        # Headers - Sprint 53: Added Ref column
        headers = ["Ref", "Account", "Type", "Issue", "Amount", "Severity", "Classification"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = "header_style"

        abnormal_balances = self.audit_result.get("abnormal_balances", [])

        # Separate material and immaterial
        material = [ab for ab in abnormal_balances if ab.get("materiality") == "material"]
        immaterial = [ab for ab in abnormal_balances if ab.get("materiality") == "immaterial"]

        row = 5

        # Material anomalies section
        if material:
            ws.cell(row=row, column=1, value=f"MATERIAL RISKS ({len(material)})")
            ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.CLAY, size=11)
            ws.merge_cells(f"A{row}:G{row}")  # Sprint 53: Extended
            row += 1

            for idx, ab in enumerate(material, start=1):
                # Sprint 53: Reference number
                ref_num = f"TB-M{idx:03d}"
                ws.cell(row=row, column=1, value=ref_num)
                ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_500, size=9)
                ws.cell(row=row, column=2, value=sanitize_csv_value(ab.get("account", "Unknown")))
                ws.cell(row=row, column=3, value=sanitize_csv_value(ab.get("type", "Unknown")))
                ws.cell(row=row, column=4, value=sanitize_csv_value(ab.get("issue", "")))
                ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=5, value=ab.get("amount", 0))
                ws.cell(row=row, column=5).style = "currency_style"
                ws.cell(row=row, column=6, value=ab.get("severity", "unknown").upper())
                ws.cell(row=row, column=7, value="MATERIAL")
                ws.cell(row=row, column=7).font = Font(bold=True, color=ExcelColors.CLAY)

                # Apply material styling
                ws.cell(row=row, column=2).border = Border(left=Side(style="thick", color=ExcelColors.CLAY))
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=ExcelColors.LIGHT_GRAY)

                row += 1

            row += 1  # Spacer

        # Immaterial anomalies section
        if immaterial:
            ws.cell(row=row, column=1, value=f"IMMATERIAL ITEMS ({len(immaterial)})")
            ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN_500, size=11)
            ws.merge_cells(f"A{row}:G{row}")  # Sprint 53: Extended
            row += 1

            for idx, ab in enumerate(immaterial, start=1):
                # Sprint 53: Reference number
                ref_num = f"TB-I{idx:03d}"
                ws.cell(row=row, column=1, value=ref_num)
                ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_500, size=9)
                ws.cell(row=row, column=2, value=sanitize_csv_value(ab.get("account", "Unknown")))
                ws.cell(row=row, column=3, value=sanitize_csv_value(ab.get("type", "Unknown")))
                ws.cell(row=row, column=4, value=sanitize_csv_value(ab.get("issue", "")))
                ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=5, value=ab.get("amount", 0))
                ws.cell(row=row, column=5).style = "currency_style"
                ws.cell(row=row, column=6, value=ab.get("severity", "unknown").upper())
                ws.cell(row=row, column=7, value="IMMATERIAL")
                ws.cell(row=row, column=7).font = Font(color=ExcelColors.OBSIDIAN_500)

                row += 1

        # If no anomalies
        if not material and not immaterial:
            ws.cell(row=row, column=1, value="No anomalies detected. Trial balance appears healthy.")
            ws.cell(row=row, column=1).font = Font(color=ExcelColors.SAGE, size=11, italic=True)
            ws.merge_cells(f"A{row}:F{row}")

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 15

    def _build_ratios_tab(self) -> None:
        """Build the Key Ratios tab with financial ratio analysis."""
        ws = self.wb.create_sheet("Key Ratios", 3)

        # Title
        ws["A1"] = "Key Financial Ratios"
        ws["A1"].style = "title_style"
        ws.merge_cells("A1:D1")

        # Subtitle
        ws["A2"] = "Based on standardized trial balance categories"
        ws["A2"].style = "subtitle_style"

        # Headers
        headers = ["Ratio", "Value", "Health Status", "Interpretation"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = "header_style"

        # Get analytics data if available
        analytics = self.audit_result.get("analytics", {})
        ratios = analytics.get("ratios", {})

        row = 5

        # Define ratio display info
        ratio_definitions = [
            ("current_ratio", "Current Ratio", "Current Assets / Current Liabilities"),
            ("quick_ratio", "Quick Ratio", "(Current Assets - Inventory) / Current Liabilities"),
            ("debt_to_equity", "Debt-to-Equity", "Total Liabilities / Total Equity"),
            ("gross_margin", "Gross Margin", "(Revenue - COGS) / Revenue"),
        ]

        for ratio_key, ratio_name, formula in ratio_definitions:
            ratio_data = ratios.get(ratio_key, {})

            ws.cell(row=row, column=1, value=ratio_name)
            ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN)

            value = ratio_data.get("value")
            if value is not None:
                # Format based on ratio type
                if ratio_key == "gross_margin":
                    ws.cell(row=row, column=2, value=value / 100 if value > 1 else value)
                    ws.cell(row=row, column=2).style = "percent_style"
                else:
                    ws.cell(row=row, column=2, value=value)
                    ws.cell(row=row, column=2).number_format = "0.00"

                # Threshold status with color
                threshold = ratio_data.get("threshold_status", "neutral")
                ws.cell(row=row, column=3, value=threshold.upper().replace("_", " "))
                if threshold == "above_threshold":
                    ws.cell(row=row, column=3).font = Font(bold=True, color=ExcelColors.SAGE)
                elif threshold == "at_threshold":
                    ws.cell(row=row, column=3).font = Font(bold=True, color=ExcelColors.OATMEAL_500)
                elif threshold == "below_threshold":
                    ws.cell(row=row, column=3).font = Font(bold=True, color=ExcelColors.CLAY)

                # Interpretation
                ws.cell(row=row, column=4, value=ratio_data.get("interpretation", ""))
                ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            else:
                ws.cell(row=row, column=2, value="N/A")
                ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_500, italic=True)
                ws.cell(row=row, column=3, value="-")
                ws.cell(row=row, column=4, value="Insufficient data to calculate")
                ws.cell(row=row, column=4).font = Font(color=ExcelColors.OBSIDIAN_500, italic=True)

            # Add formula note below
            row += 1
            ws.cell(row=row, column=1, value=f"  Formula: {formula}")
            ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)
            ws.merge_cells(f"A{row}:D{row}")

            row += 2  # Spacer

        # Category totals section
        row += 2
        ws.cell(row=row, column=1, value="Category Totals (Used in Calculations)")
        ws.cell(row=row, column=1).style = "title_style"
        ws.merge_cells(f"A{row}:D{row}")

        row += 2
        category_totals = analytics.get("category_totals", {})
        categories = [
            ("total_assets", "Total Assets"),
            ("current_assets", "Current Assets"),
            ("inventory", "Inventory"),
            ("total_liabilities", "Total Liabilities"),
            ("current_liabilities", "Current Liabilities"),
            ("total_equity", "Total Equity"),
            ("total_revenue", "Total Revenue"),
            ("cost_of_goods_sold", "Cost of Goods Sold"),
            ("total_expenses", "Total Expenses"),
        ]

        for key, label in categories:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600)
            ws.cell(row=row, column=2, value=category_totals.get(key, 0))
            ws.cell(row=row, column=2).style = "currency_style"
            row += 1

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 50


def generate_financial_statements_excel(
    statements,
    prepared_by: Optional[str] = None,
    reviewed_by: Optional[str] = None,
    workpaper_date: Optional[str] = None,
    include_signoff: bool = False,
) -> bytes:
    """
    Generate an Excel workbook with Balance Sheet and Income Statement tabs.

    Sprint 71: Financial Statements Excel using Oat & Obsidian theme.

    Args:
        statements: FinancialStatements dataclass from FinancialStatementBuilder
        prepared_by: Name of preparer (deprecated — ignored unless include_signoff=True)
        reviewed_by: Name of reviewer (deprecated — ignored unless include_signoff=True)
        workpaper_date: Date for workpaper signoff (deprecated — ignored unless include_signoff=True)
        include_signoff: If True, render signoff section (default False since Sprint 7)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Register styles
    for style_fn in [
        create_header_style,
        create_title_style,
        create_subtitle_style,
        create_currency_style,
        create_balanced_style,
        create_unbalanced_style,
    ]:
        try:
            wb.add_named_style(style_fn())
        except ValueError:
            pass

    def _write_statement_sheet(ws: Worksheet, title: str, line_items, sheet_index: int) -> int:
        """Write a financial statement to a worksheet."""
        # Title
        ws["A1"] = title
        ws["A1"].style = "title_style"
        ws.merge_cells("A1:C1")

        # Entity name
        entity = statements.entity_name or ""
        if entity:
            ws["A2"] = entity
            ws["A2"].style = "subtitle_style"
            ws.merge_cells("A2:C2")

        # Period
        if statements.period_end:
            ws["A3"] = f"Period Ending: {statements.period_end}"
            ws["A3"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        # Timestamp
        ws["A4"] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A4"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        row = 6
        for item in line_items:
            # Column A: label (with indentation)
            indent = "    " * item.indent_level
            label = f"{indent}{item.label}"

            ws.cell(row=row, column=1, value=label)

            # Style based on line type
            if item.is_total:
                ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)
                ws.cell(row=row, column=2).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)
                ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=2).border = Border(bottom=Side(style="double", color=ExcelColors.OBSIDIAN))
                ws.cell(row=row, column=2, value=item.amount)
            elif item.is_subtotal:
                ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN_600, size=10)
                ws.cell(row=row, column=2).font = Font(bold=True, color=ExcelColors.OBSIDIAN_600, size=10)
                ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=2).border = Border(top=Side(style="thin", color=ExcelColors.OBSIDIAN_500))
                ws.cell(row=row, column=2, value=item.amount)
            elif item.indent_level == 0 and not item.lead_sheet_ref:
                # Section header — no amount
                ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=10)
                ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ExcelColors.OATMEAL)
            else:
                # Regular line item
                ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                ws.cell(row=row, column=2, value=item.amount)
                ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
                ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)

            # Column C: lead sheet reference
            if item.lead_sheet_ref:
                ws.cell(row=row, column=3, value=item.lead_sheet_ref)
                ws.cell(row=row, column=3).font = Font(color=ExcelColors.OBSIDIAN_500, size=8, italic=True)
                ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

            row += 1

        return row

    # ── Balance Sheet tab ──
    bs_ws = wb.create_sheet("Balance Sheet", 0)
    end_row = _write_statement_sheet(bs_ws, "Balance Sheet", statements.balance_sheet, 0)

    # Balance verification row
    end_row += 1
    if statements.is_balanced:
        bs_ws.cell(row=end_row, column=1, value="✓ BALANCED")
        bs_ws.cell(row=end_row, column=1).style = "balanced_style"
    else:
        bs_ws.cell(row=end_row, column=1, value=f"⚠ OUT OF BALANCE (${statements.balance_difference:,.2f})")
        bs_ws.cell(row=end_row, column=1).style = "unbalanced_style"
    bs_ws.merge_cells(f"A{end_row}:C{end_row}")

    # Workpaper signoff (deprecated Sprint 7 — gated by include_signoff)
    if include_signoff and (prepared_by or reviewed_by):
        end_row += 2
        wp_date = workpaper_date or datetime.now().strftime("%Y-%m-%d")
        bs_ws.cell(row=end_row, column=1, value="Workpaper Sign-Off")
        bs_ws.cell(row=end_row, column=1).style = "title_style"
        end_row += 1
        if prepared_by:
            bs_ws.cell(row=end_row, column=1, value="Prepared By:")
            bs_ws.cell(row=end_row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            bs_ws.cell(row=end_row, column=2, value=prepared_by)
            bs_ws.cell(row=end_row, column=3, value=wp_date)
            end_row += 1
        if reviewed_by:
            bs_ws.cell(row=end_row, column=1, value="Reviewed By:")
            bs_ws.cell(row=end_row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            bs_ws.cell(row=end_row, column=2, value=reviewed_by)
            bs_ws.cell(row=end_row, column=3, value=wp_date)

    bs_ws.column_dimensions["A"].width = 40
    bs_ws.column_dimensions["B"].width = 18
    bs_ws.column_dimensions["C"].width = 8

    # ── Income Statement tab ──
    is_ws = wb.create_sheet("Income Statement", 1)
    _write_statement_sheet(is_ws, "Income Statement", statements.income_statement, 1)

    is_ws.column_dimensions["A"].width = 40
    is_ws.column_dimensions["B"].width = 18
    is_ws.column_dimensions["C"].width = 8

    # ── Cash Flow Statement tab (Sprint 84) ──
    if statements.cash_flow_statement is not None:
        cf = statements.cash_flow_statement
        cf_ws = wb.create_sheet("Cash Flow Statement", 2)

        # Title
        cf_ws["A1"] = "Cash Flow Statement (Indirect Method)"
        cf_ws["A1"].style = "title_style"
        cf_ws.merge_cells("A1:C1")

        entity = statements.entity_name or ""
        if entity:
            cf_ws["A2"] = entity
            cf_ws["A2"].style = "subtitle_style"
            cf_ws.merge_cells("A2:C2")

        if statements.period_end:
            cf_ws["A3"] = f"Period Ending: {statements.period_end}"
            cf_ws["A3"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        cf_ws["A4"] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
        cf_ws["A4"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        row = 6
        for section in [cf.operating, cf.investing, cf.financing]:
            # Section header
            cf_ws.cell(row=row, column=1, value=section.label)
            cf_ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=10)
            cf_ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ExcelColors.OATMEAL)
            row += 1

            # Line items
            for item in section.items:
                indent = "    " * item.indent_level
                cf_ws.cell(row=row, column=1, value=f"{indent}{item.label}")
                cf_ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                cf_ws.cell(row=row, column=2, value=item.amount)
                cf_ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
                cf_ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                if item.source:
                    cf_ws.cell(row=row, column=3, value=item.source)
                    cf_ws.cell(row=row, column=3).font = Font(color=ExcelColors.OBSIDIAN_500, size=8, italic=True)
                    cf_ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                row += 1

            # Section subtotal
            cf_ws.cell(row=row, column=1, value=f"    Net {section.label}")
            cf_ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN_600, size=10)
            cf_ws.cell(row=row, column=2, value=section.subtotal)
            cf_ws.cell(row=row, column=2).font = Font(bold=True, color=ExcelColors.OBSIDIAN_600, size=10)
            cf_ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
            cf_ws.cell(row=row, column=2).border = Border(top=Side(style="thin", color=ExcelColors.OBSIDIAN_500))
            row += 2

        # Net Change in Cash
        cf_ws.cell(row=row, column=1, value="NET CHANGE IN CASH")
        cf_ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)
        cf_ws.cell(row=row, column=2, value=cf.net_change)
        cf_ws.cell(row=row, column=2).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)
        cf_ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        cf_ws.cell(row=row, column=2).border = Border(bottom=Side(style="double", color=ExcelColors.OBSIDIAN))
        row += 2

        # Reconciliation
        if cf.has_prior_period:
            cf_ws.cell(row=row, column=1, value="Reconciliation")
            cf_ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=10)
            cf_ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=ExcelColors.OATMEAL)
            row += 1

            cf_ws.cell(row=row, column=1, value="    Beginning Cash")
            cf_ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            cf_ws.cell(row=row, column=2, value=cf.beginning_cash)
            cf_ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
            cf_ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            row += 1

            cf_ws.cell(row=row, column=1, value="    Net Change in Cash")
            cf_ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            cf_ws.cell(row=row, column=2, value=cf.net_change)
            cf_ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
            cf_ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
            row += 1

            cf_ws.cell(row=row, column=1, value="ENDING CASH")
            cf_ws.cell(row=row, column=1).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)
            cf_ws.cell(row=row, column=2, value=cf.ending_cash)
            cf_ws.cell(row=row, column=2).font = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)
            cf_ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
            cf_ws.cell(row=row, column=2).border = Border(bottom=Side(style="double", color=ExcelColors.OBSIDIAN))
            row += 2

            # Reconciliation status
            if cf.is_reconciled:
                cf_ws.cell(row=row, column=1, value="✓ RECONCILED")
                cf_ws.cell(row=row, column=1).style = "balanced_style"
            else:
                cf_ws.cell(row=row, column=1, value=f"⚠ UNRECONCILED (${cf.reconciliation_difference:,.2f})")
                cf_ws.cell(row=row, column=1).style = "unbalanced_style"
            cf_ws.merge_cells(f"A{row}:C{row}")
            row += 1

        # Notes
        if cf.notes:
            row += 1
            for note in cf.notes:
                cf_ws.cell(row=row, column=1, value=f"Note: {note}")
                cf_ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)
                row += 1

        cf_ws.column_dimensions["A"].width = 40
        cf_ws.column_dimensions["B"].width = 18
        cf_ws.column_dimensions["C"].width = 8

    # ── Mapping Trace tab (Sprint 284) ──
    if statements.mapping_trace:
        mt_ws = wb.create_sheet("Mapping Trace", 3)

        # Title
        mt_ws["A1"] = "Account-to-Statement Mapping Trace"
        mt_ws["A1"].style = "title_style"
        mt_ws.merge_cells("A1:H1")

        entity = statements.entity_name or ""
        if entity:
            mt_ws["A2"] = entity
            mt_ws["A2"].style = "subtitle_style"
            mt_ws.merge_cells("A2:H2")

        if statements.period_end:
            mt_ws["A3"] = f"Period Ending: {statements.period_end}"
            mt_ws["A3"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        mt_ws["A4"] = f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}"
        mt_ws["A4"].font = Font(color=ExcelColors.OBSIDIAN_500, size=9, italic=True)

        # Header row
        row = 6
        headers = ["Statement", "Line Item", "Ref", "Account", "Debit", "Credit", "Net", "Tied"]
        for col_idx, header in enumerate(headers, 1):
            cell = mt_ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True, color=ExcelColors.OATMEAL, size=10)
            cell.fill = PatternFill("solid", fgColor=ExcelColors.OBSIDIAN)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        for entry in statements.mapping_trace:
            first_row = True
            if entry.account_count > 0 and entry.accounts:
                for acct in entry.accounts:
                    if first_row:
                        mt_ws.cell(row=row, column=1, value=entry.statement)
                        mt_ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                        mt_ws.cell(row=row, column=2, value=entry.line_label)
                        mt_ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                        mt_ws.cell(row=row, column=3, value=entry.lead_sheet_ref)
                        mt_ws.cell(row=row, column=3).font = Font(color=ExcelColors.OBSIDIAN_500, size=9)
                        mt_ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                        first_row = False

                    mt_ws.cell(row=row, column=4, value=acct.account_name)
                    mt_ws.cell(row=row, column=4).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                    mt_ws.cell(row=row, column=5, value=acct.debit)
                    mt_ws.cell(row=row, column=5).number_format = '"$"#,##0.00'
                    mt_ws.cell(row=row, column=5).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                    mt_ws.cell(row=row, column=6, value=acct.credit)
                    mt_ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
                    mt_ws.cell(row=row, column=6).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                    mt_ws.cell(row=row, column=7, value=acct.net_balance)
                    mt_ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
                    mt_ws.cell(row=row, column=7).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                    row += 1

                # Subtotal row
                mt_ws.cell(row=row, column=2, value=f"Subtotal: {entry.line_label}")
                mt_ws.cell(row=row, column=2).font = Font(bold=True, color=ExcelColors.OBSIDIAN_600, size=10)
                mt_ws.cell(row=row, column=7, value=entry.statement_amount)
                mt_ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
                mt_ws.cell(row=row, column=7).font = Font(bold=True, color=ExcelColors.OBSIDIAN_600, size=10)
                mt_ws.cell(row=row, column=7).border = Border(top=Side(style="thin", color=ExcelColors.OBSIDIAN_500))

                # Tied indicator
                if entry.is_tied:
                    mt_ws.cell(row=row, column=8, value="✓")
                    mt_ws.cell(row=row, column=8).font = Font(color=ExcelColors.SAGE, size=10, bold=True)
                else:
                    mt_ws.cell(row=row, column=8, value="⚠")
                    mt_ws.cell(row=row, column=8).font = Font(color=ExcelColors.CLAY, size=10, bold=True)
                mt_ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")
                row += 1
            else:
                # Empty lead sheet — single row
                mt_ws.cell(row=row, column=1, value=entry.statement)
                mt_ws.cell(row=row, column=1).font = Font(color=ExcelColors.OBSIDIAN_600, size=10)
                mt_ws.cell(row=row, column=2, value=entry.line_label)
                mt_ws.cell(row=row, column=2).font = Font(color=ExcelColors.OBSIDIAN_500, size=10, italic=True)
                mt_ws.cell(row=row, column=3, value=entry.lead_sheet_ref)
                mt_ws.cell(row=row, column=3).font = Font(color=ExcelColors.OBSIDIAN_500, size=9)
                mt_ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
                mt_ws.cell(row=row, column=7, value=0.0)
                mt_ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
                mt_ws.cell(row=row, column=8, value="✓")
                mt_ws.cell(row=row, column=8).font = Font(color=ExcelColors.SAGE, size=10, bold=True)
                mt_ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")
                row += 1

        mt_ws.column_dimensions["A"].width = 16
        mt_ws.column_dimensions["B"].width = 30
        mt_ws.column_dimensions["C"].width = 6
        mt_ws.column_dimensions["D"].width = 35
        mt_ws.column_dimensions["E"].width = 15
        mt_ws.column_dimensions["F"].width = 15
        mt_ws.column_dimensions["G"].width = 15
        mt_ws.column_dimensions["H"].width = 8

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    buf.close()

    log_secure_operation(
        "financial_statements_excel_complete", f"Financial statements Excel generated: {len(excel_bytes)} bytes"
    )

    return excel_bytes


def generate_workpaper(
    audit_result: dict[str, Any],
    filename: str = "workpaper",
    prepared_by: Optional[str] = None,
    reviewed_by: Optional[str] = None,
    workpaper_date: Optional[str] = None,
    include_signoff: bool = False,
) -> bytes:
    """
    Generate an Excel workpaper from audit results.

    Sprint 53: Added workpaper fields for professional documentation.
    Sprint 7: Signoff deprecated — gated by include_signoff (default False).

    Args:
        audit_result: The audit result dictionary
        filename: Base filename for the workpaper
        prepared_by: Name of preparer (deprecated — ignored unless include_signoff=True)
        reviewed_by: Name of reviewer (deprecated — ignored unless include_signoff=True)
        workpaper_date: Date for workpaper signoff (deprecated — ignored unless include_signoff=True)
        include_signoff: If True, render signoff section (default False since Sprint 7)
    """
    generator = PaciolusWorkpaperGenerator(
        audit_result,
        filename,
        prepared_by=prepared_by,
        reviewed_by=reviewed_by,
        workpaper_date=workpaper_date,
        include_signoff=include_signoff,
    )
    return generator.generate()
