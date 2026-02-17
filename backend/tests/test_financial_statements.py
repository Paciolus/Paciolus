"""
Tests for Financial Statement Builder, PDF, and Excel export.
Sprint 71: Financial Statements — Backend Builder + Export
"""

from io import BytesIO

import pytest

from financial_statement_builder import (
    FinancialStatementBuilder,
)

# =============================================================================
# FIXTURES
# =============================================================================

@pytest.fixture
def sample_lead_sheet_grouping():
    """
    Balanced lead sheet grouping fixture.

    Assets: A=$50K, B=$30K, E=$45K, F=$5K → Total Assets = $130K
    Liabilities: G=-$25K, I=-$30K → Total Liabilities = $55K
    Equity: K=-$75K → Total Equity = $75K
    Total L+E = $130K (balanced)

    Revenue: L=-$150K, COGS: M=$60K, OpEx: N=$40K, Other: O=-$5K
    Net Income = $150K - $60K - $40K + $5K = $55K
    """
    return {
        "summaries": [
            {
                "lead_sheet": "A",
                "name": "Cash and Cash Equivalents",
                "category": "Current Assets",
                "total_debit": 50000.0,
                "total_credit": 0.0,
                "net_balance": 50000.0,
                "account_count": 2,
                "accounts": [],
            },
            {
                "lead_sheet": "B",
                "name": "Receivables",
                "category": "Current Assets",
                "total_debit": 30000.0,
                "total_credit": 0.0,
                "net_balance": 30000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "E",
                "name": "Property, Plant & Equipment",
                "category": "Non-Current Assets",
                "total_debit": 45000.0,
                "total_credit": 0.0,
                "net_balance": 45000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "F",
                "name": "Other Assets & Intangibles",
                "category": "Non-Current Assets",
                "total_debit": 5000.0,
                "total_credit": 0.0,
                "net_balance": 5000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "G",
                "name": "Accounts Payable & Accrued Liabilities",
                "category": "Current Liabilities",
                "total_debit": 0.0,
                "total_credit": 25000.0,
                "net_balance": -25000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "I",
                "name": "Long-term Debt",
                "category": "Non-Current Liabilities",
                "total_debit": 0.0,
                "total_credit": 30000.0,
                "net_balance": -30000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "K",
                "name": "Stockholders' Equity",
                "category": "Equity",
                "total_debit": 0.0,
                "total_credit": 75000.0,
                "net_balance": -75000.0,
                "account_count": 2,
                "accounts": [],
            },
            {
                "lead_sheet": "L",
                "name": "Revenue",
                "category": "Revenue",
                "total_debit": 0.0,
                "total_credit": 150000.0,
                "net_balance": -150000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "M",
                "name": "Cost of Goods Sold",
                "category": "Cost of Sales",
                "total_debit": 60000.0,
                "total_credit": 0.0,
                "net_balance": 60000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "N",
                "name": "Operating Expenses",
                "category": "Operating Expenses",
                "total_debit": 40000.0,
                "total_credit": 0.0,
                "net_balance": 40000.0,
                "account_count": 3,
                "accounts": [],
            },
            {
                "lead_sheet": "O",
                "name": "Other Income & Expense",
                "category": "Other",
                "total_debit": 0.0,
                "total_credit": 5000.0,
                "net_balance": -5000.0,
                "account_count": 1,
                "accounts": [],
            },
        ],
        "total_debits": 230000.0,
        "total_credits": 285000.0,
        "unclassified_count": 0,
    }


@pytest.fixture
def unbalanced_lead_sheet_grouping():
    """Lead sheet grouping that does NOT balance (assets != L+E)."""
    return {
        "summaries": [
            {
                "lead_sheet": "A",
                "name": "Cash",
                "category": "Current Assets",
                "total_debit": 100000.0,
                "total_credit": 0.0,
                "net_balance": 100000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "K",
                "name": "Equity",
                "category": "Equity",
                "total_debit": 0.0,
                "total_credit": 80000.0,
                "net_balance": -80000.0,
                "account_count": 1,
                "accounts": [],
            },
        ],
        "total_debits": 100000.0,
        "total_credits": 80000.0,
        "unclassified_count": 0,
    }


@pytest.fixture
def other_expense_grouping():
    """Lead sheet grouping where O has a debit (expense) balance."""
    return {
        "summaries": [
            {
                "lead_sheet": "L",
                "name": "Revenue",
                "category": "Revenue",
                "total_debit": 0.0,
                "total_credit": 100000.0,
                "net_balance": -100000.0,
                "account_count": 1,
                "accounts": [],
            },
            {
                "lead_sheet": "O",
                "name": "Other Income & Expense",
                "category": "Other",
                "total_debit": 8000.0,
                "total_credit": 0.0,
                "net_balance": 8000.0,
                "account_count": 1,
                "accounts": [],
            },
        ],
        "total_debits": 8000.0,
        "total_credits": 100000.0,
        "unclassified_count": 0,
    }


# =============================================================================
# TestFinancialStatementBuilder
# =============================================================================

class TestFinancialStatementBuilder:
    """Tests for the FinancialStatementBuilder class."""

    def test_balance_sheet_current_assets(self, sample_lead_sheet_grouping):
        """A-D grouped as current assets, positive amounts."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        # Find current asset line items
        bs_items = {item.lead_sheet_ref: item for item in result.balance_sheet if item.lead_sheet_ref}
        assert bs_items["A"].amount == 50000.0
        assert bs_items["B"].amount == 30000.0

    def test_balance_sheet_noncurrent_assets(self, sample_lead_sheet_grouping):
        """E-F grouped as non-current assets."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        bs_items = {item.lead_sheet_ref: item for item in result.balance_sheet if item.lead_sheet_ref}
        assert bs_items["E"].amount == 45000.0
        assert bs_items["F"].amount == 5000.0

    def test_balance_sheet_current_liabilities(self, sample_lead_sheet_grouping):
        """G-H grouped, sign-flipped to positive."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        bs_items = {item.lead_sheet_ref: item for item in result.balance_sheet if item.lead_sheet_ref}
        assert bs_items["G"].amount == 25000.0  # Flipped from -25000

    def test_balance_sheet_noncurrent_liabilities(self, sample_lead_sheet_grouping):
        """I-J grouped, sign-flipped."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        bs_items = {item.lead_sheet_ref: item for item in result.balance_sheet if item.lead_sheet_ref}
        assert bs_items["I"].amount == 30000.0  # Flipped from -30000

    def test_balance_sheet_equity(self, sample_lead_sheet_grouping):
        """K sign-flipped."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        bs_items = {item.lead_sheet_ref: item for item in result.balance_sheet if item.lead_sheet_ref}
        assert bs_items["K"].amount == 75000.0  # Flipped from -75000

    def test_balance_sheet_total_assets(self, sample_lead_sheet_grouping):
        """Total assets = current + non-current."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        assert result.total_assets == 130000.0  # 50+30+45+5

    def test_balance_sheet_balanced(self, sample_lead_sheet_grouping):
        """is_balanced=True when assets = L+E."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        assert result.is_balanced is True
        assert abs(result.balance_difference) < 0.01

    def test_balance_sheet_unbalanced(self, unbalanced_lead_sheet_grouping):
        """is_balanced=False, difference reported."""
        builder = FinancialStatementBuilder(unbalanced_lead_sheet_grouping)
        result = builder.build()

        assert result.is_balanced is False
        assert result.balance_difference == 20000.0  # 100K assets - 80K equity

    def test_income_statement_revenue(self, sample_lead_sheet_grouping):
        """L sign-flipped to positive."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        assert result.total_revenue == 150000.0

    def test_income_statement_cogs(self, sample_lead_sheet_grouping):
        """M kept positive (debit balance)."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        is_items = {item.lead_sheet_ref: item for item in result.income_statement if item.lead_sheet_ref}
        assert is_items["M"].amount == 60000.0

    def test_income_statement_gross_profit(self, sample_lead_sheet_grouping):
        """Gross profit = revenue - COGS."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        assert result.gross_profit == 90000.0  # 150K - 60K

    def test_income_statement_operating_expenses(self, sample_lead_sheet_grouping):
        """N kept positive (debit balance)."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        is_items = {item.lead_sheet_ref: item for item in result.income_statement if item.lead_sheet_ref}
        assert is_items["N"].amount == 40000.0

    def test_income_statement_operating_income(self, sample_lead_sheet_grouping):
        """Operating income = gross profit - opex."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        assert result.operating_income == 50000.0  # 90K - 40K

    def test_income_statement_other_income(self, sample_lead_sheet_grouping):
        """O with credit balance → positive (income)."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        is_items = {item.lead_sheet_ref: item for item in result.income_statement if item.lead_sheet_ref}
        assert is_items["O"].amount == 5000.0  # Flipped from -5000

    def test_income_statement_other_expense(self, other_expense_grouping):
        """O with debit balance → negative (expense)."""
        builder = FinancialStatementBuilder(other_expense_grouping)
        result = builder.build()

        is_items = {item.lead_sheet_ref: item for item in result.income_statement if item.lead_sheet_ref}
        assert is_items["O"].amount == -8000.0  # Flipped from +8000 debit

    def test_income_statement_net_income(self, sample_lead_sheet_grouping):
        """Net income = operating + other."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()

        assert result.net_income == 55000.0  # 50K + 5K

    def test_missing_lead_sheets(self):
        """Graceful zero for absent categories."""
        grouping = {
            "summaries": [
                {
                    "lead_sheet": "A",
                    "name": "Cash",
                    "category": "Current Assets",
                    "total_debit": 10000.0,
                    "total_credit": 0.0,
                    "net_balance": 10000.0,
                    "account_count": 1,
                    "accounts": [],
                },
            ],
            "total_debits": 10000.0,
            "total_credits": 0.0,
            "unclassified_count": 0,
        }
        builder = FinancialStatementBuilder(grouping)
        result = builder.build()

        assert result.total_assets == 10000.0
        assert result.total_liabilities == 0.0
        assert result.total_equity == 0.0
        assert result.total_revenue == 0.0
        assert result.net_income == 0.0

    def test_empty_grouping(self):
        """Handles empty summaries list."""
        grouping = {
            "summaries": [],
            "total_debits": 0.0,
            "total_credits": 0.0,
            "unclassified_count": 0,
        }
        builder = FinancialStatementBuilder(grouping)
        result = builder.build()

        assert result.total_assets == 0.0
        assert result.is_balanced is True
        assert len(result.balance_sheet) > 0  # Structure still generated
        assert len(result.income_statement) > 0

    def test_to_dict_serialization(self, sample_lead_sheet_grouping):
        """Round-trip serializable."""
        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        result = builder.build()
        d = result.to_dict()

        assert isinstance(d, dict)
        assert "balance_sheet" in d
        assert "income_statement" in d
        assert isinstance(d["balance_sheet"], list)
        assert isinstance(d["income_statement"], list)
        assert d["total_assets"] == 130000.0
        assert d["is_balanced"] is True
        assert d["net_income"] == 55000.0

        # Verify each line item is a dict
        for item in d["balance_sheet"]:
            assert "label" in item
            assert "amount" in item
            assert "indent_level" in item

    def test_zero_revenue_no_division_error(self):
        """Gross margin calculation handles zero revenue."""
        grouping = {
            "summaries": [
                {
                    "lead_sheet": "M",
                    "name": "COGS",
                    "category": "Cost of Sales",
                    "total_debit": 5000.0,
                    "total_credit": 0.0,
                    "net_balance": 5000.0,
                    "account_count": 1,
                    "accounts": [],
                },
            ],
            "total_debits": 5000.0,
            "total_credits": 0.0,
            "unclassified_count": 0,
        }
        builder = FinancialStatementBuilder(grouping)
        result = builder.build()

        assert result.total_revenue == 0.0
        assert result.gross_profit == -5000.0
        assert result.net_income == -5000.0

    def test_entity_name_and_period(self, sample_lead_sheet_grouping):
        """Metadata fields passed through."""
        builder = FinancialStatementBuilder(
            sample_lead_sheet_grouping,
            entity_name="Acme Corp",
            period_end="2025-12-31",
        )
        result = builder.build()

        assert result.entity_name == "Acme Corp"
        assert result.period_end == "2025-12-31"

        d = result.to_dict()
        assert d["entity_name"] == "Acme Corp"
        assert d["period_end"] == "2025-12-31"


# =============================================================================
# TestFinancialStatementPDF
# =============================================================================

class TestFinancialStatementPDF:
    """Tests for financial statement PDF generation."""

    def test_pdf_bytes_nonempty(self, sample_lead_sheet_grouping):
        """Generate PDF, assert non-empty bytes."""
        from pdf_generator import generate_financial_statements_pdf

        builder = FinancialStatementBuilder(sample_lead_sheet_grouping, entity_name="Test Corp")
        statements = builder.build()

        pdf_bytes = generate_financial_statements_pdf(statements)
        assert len(pdf_bytes) > 0

    def test_pdf_bytes_type(self, sample_lead_sheet_grouping):
        """Assert PDF output is bytes."""
        from pdf_generator import generate_financial_statements_pdf

        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        statements = builder.build()

        pdf_bytes = generate_financial_statements_pdf(statements)
        assert isinstance(pdf_bytes, bytes)

    def test_pdf_with_workpaper_fields(self, sample_lead_sheet_grouping):
        """PDF generates with workpaper signoff fields."""
        from pdf_generator import generate_financial_statements_pdf

        builder = FinancialStatementBuilder(sample_lead_sheet_grouping, entity_name="Acme Corp")
        statements = builder.build()

        pdf_bytes = generate_financial_statements_pdf(
            statements,
            prepared_by="John Doe",
            reviewed_by="Jane Smith",
        )
        assert len(pdf_bytes) > 100


# =============================================================================
# TestFinancialStatementExcel
# =============================================================================

class TestFinancialStatementExcel:
    """Tests for financial statement Excel generation."""

    def test_excel_bytes_nonempty(self, sample_lead_sheet_grouping):
        """Generate Excel, assert non-empty bytes."""
        from excel_generator import generate_financial_statements_excel

        builder = FinancialStatementBuilder(sample_lead_sheet_grouping, entity_name="Test Corp")
        statements = builder.build()

        excel_bytes = generate_financial_statements_excel(statements)
        assert len(excel_bytes) > 0

    def test_excel_has_balance_sheet_tab(self, sample_lead_sheet_grouping):
        """Load workbook, check sheet names."""
        from openpyxl import load_workbook

        from excel_generator import generate_financial_statements_excel

        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        statements = builder.build()

        excel_bytes = generate_financial_statements_excel(statements)
        wb = load_workbook(BytesIO(excel_bytes))
        assert "Balance Sheet" in wb.sheetnames

    def test_excel_has_income_statement_tab(self, sample_lead_sheet_grouping):
        """Load workbook, check sheet names."""
        from openpyxl import load_workbook

        from excel_generator import generate_financial_statements_excel

        builder = FinancialStatementBuilder(sample_lead_sheet_grouping)
        statements = builder.build()

        excel_bytes = generate_financial_statements_excel(statements)
        wb = load_workbook(BytesIO(excel_bytes))
        assert "Income Statement" in wb.sheetnames


# =============================================================================
# MAPPING TRACE FIXTURES
# =============================================================================

@pytest.fixture
def grouping_with_accounts():
    """
    Lead sheet grouping with populated accounts arrays for mapping trace tests.

    A: 2 accounts (Cash at Bank $30K, Petty Cash $20K) → $50K
    B: 2 accounts (Trade Receivables $20K, Other Receivables $10K) → $30K
    G: 2 accounts (Trade Payables $15K cr, Accrued Expenses $10K cr) → -$25K
    K: 2 accounts (Common Stock $50K cr, Retained Earnings $25K cr) → -$75K
    L: 2 accounts (Product Revenue $100K cr, Service Revenue $50K cr) → -$150K
    N: 3 accounts (Salaries $20K, Rent $10K, Depreciation $10K) → $40K
    """
    return {
        "summaries": [
            {
                "lead_sheet": "A",
                "name": "Cash and Cash Equivalents",
                "category": "Current Assets",
                "total_debit": 50000.0,
                "total_credit": 0.0,
                "net_balance": 50000.0,
                "account_count": 2,
                "accounts": [
                    {"account": "Cash at Bank", "debit": 30000.0, "credit": 0.0, "confidence": 0.95, "matched_keywords": ["cash", "bank"]},
                    {"account": "Petty Cash", "debit": 20000.0, "credit": 0.0, "confidence": 0.90, "matched_keywords": ["cash"]},
                ],
            },
            {
                "lead_sheet": "B",
                "name": "Receivables",
                "category": "Current Assets",
                "total_debit": 30000.0,
                "total_credit": 0.0,
                "net_balance": 30000.0,
                "account_count": 2,
                "accounts": [
                    {"account": "Trade Receivables", "debit": 20000.0, "credit": 0.0, "confidence": 0.85, "matched_keywords": ["receivable"]},
                    {"account": "Other Receivables", "debit": 10000.0, "credit": 0.0, "confidence": 0.80, "matched_keywords": ["receivable"]},
                ],
            },
            {
                "lead_sheet": "G",
                "name": "AP & Accrued Liabilities",
                "category": "Current Liabilities",
                "total_debit": 0.0,
                "total_credit": 25000.0,
                "net_balance": -25000.0,
                "account_count": 2,
                "accounts": [
                    {"account": "Trade Payables", "debit": 0.0, "credit": 15000.0, "confidence": 0.90, "matched_keywords": ["payable"]},
                    {"account": "Accrued Expenses", "debit": 0.0, "credit": 10000.0, "confidence": 0.85, "matched_keywords": ["accrued"]},
                ],
            },
            {
                "lead_sheet": "K",
                "name": "Stockholders' Equity",
                "category": "Equity",
                "total_debit": 0.0,
                "total_credit": 75000.0,
                "net_balance": -75000.0,
                "account_count": 2,
                "accounts": [
                    {"account": "Common Stock", "debit": 0.0, "credit": 50000.0, "confidence": 1.0, "matched_keywords": ["stock"]},
                    {"account": "Retained Earnings", "debit": 0.0, "credit": 25000.0, "confidence": 1.0, "matched_keywords": ["retained"]},
                ],
            },
            {
                "lead_sheet": "L",
                "name": "Revenue",
                "category": "Revenue",
                "total_debit": 0.0,
                "total_credit": 150000.0,
                "net_balance": -150000.0,
                "account_count": 2,
                "accounts": [
                    {"account": "Product Revenue", "debit": 0.0, "credit": 100000.0, "confidence": 0.95, "matched_keywords": ["revenue"]},
                    {"account": "Service Revenue", "debit": 0.0, "credit": 50000.0, "confidence": 0.90, "matched_keywords": ["revenue"]},
                ],
            },
            {
                "lead_sheet": "N",
                "name": "Operating Expenses",
                "category": "Operating Expenses",
                "total_debit": 40000.0,
                "total_credit": 0.0,
                "net_balance": 40000.0,
                "account_count": 3,
                "accounts": [
                    {"account": "Salaries", "debit": 20000.0, "credit": 0.0, "confidence": 0.90, "matched_keywords": ["salary"]},
                    {"account": "Rent Expense", "debit": 10000.0, "credit": 0.0, "confidence": 0.85, "matched_keywords": ["rent"]},
                    {"name": "Depreciation Expense", "debit": 10000.0, "credit": 0.0, "confidence": 0.95, "matched_keywords": ["depreciation"]},
                ],
            },
        ],
        "total_debits": 120000.0,
        "total_credits": 250000.0,
        "unclassified_count": 0,
    }


# =============================================================================
# TestMappingTrace
# =============================================================================

class TestMappingTrace:
    """Tests for Account-to-Statement Mapping Trace (Sprint 284)."""

    def test_mapping_trace_populated(self, grouping_with_accounts):
        """Verify non-empty list with entries for populated lead sheets."""
        builder = FinancialStatementBuilder(grouping_with_accounts)
        result = builder.build()

        assert len(result.mapping_trace) == 15  # A-O
        refs = {e.lead_sheet_ref for e in result.mapping_trace}
        for letter in ["A", "B", "G", "K", "L", "N"]:
            assert letter in refs

    def test_mapping_trace_accounts_match(self, grouping_with_accounts):
        """Verify account names and amounts in trace entries."""
        builder = FinancialStatementBuilder(grouping_with_accounts)
        result = builder.build()

        entry_map = {e.lead_sheet_ref: e for e in result.mapping_trace}

        # Cash (A): 2 accounts, $30K + $20K
        a_entry = entry_map["A"]
        assert a_entry.account_count == 2
        names = [acct.account_name for acct in a_entry.accounts]
        assert "Cash at Bank" in names
        assert "Petty Cash" in names
        assert a_entry.accounts[0].debit == 30000.0
        assert a_entry.accounts[1].debit == 20000.0

    def test_mapping_trace_tie_out(self, grouping_with_accounts):
        """Verify is_tied=True when account sums match net_balance."""
        builder = FinancialStatementBuilder(grouping_with_accounts)
        result = builder.build()

        entry_map = {e.lead_sheet_ref: e for e in result.mapping_trace}
        for ref in ["A", "B", "G", "K", "L", "N"]:
            entry = entry_map[ref]
            assert entry.is_tied is True, f"Entry {ref} should be tied"
            assert entry.tie_difference < 0.01

    def test_mapping_trace_empty_lead_sheet(self, grouping_with_accounts):
        """Letters with no accounts have account_count=0, is_tied=True."""
        builder = FinancialStatementBuilder(grouping_with_accounts)
        result = builder.build()

        entry_map = {e.lead_sheet_ref: e for e in result.mapping_trace}
        # C (Inventory) not in fixture → empty
        c_entry = entry_map["C"]
        assert c_entry.account_count == 0
        assert c_entry.is_tied is True
        assert c_entry.tie_difference == 0.0

    def test_mapping_trace_in_to_dict(self, grouping_with_accounts):
        """Verify serialization includes mapping_trace key."""
        builder = FinancialStatementBuilder(grouping_with_accounts)
        result = builder.build()
        d = result.to_dict()

        assert "mapping_trace" in d
        assert isinstance(d["mapping_trace"], list)
        assert len(d["mapping_trace"]) == 15

        # Verify first entry structure
        first = d["mapping_trace"][0]
        assert "statement" in first
        assert "line_label" in first
        assert "lead_sheet_ref" in first
        assert "accounts" in first
        assert "is_tied" in first

    def test_mapping_trace_account_name_fallback(self, grouping_with_accounts):
        """Accounts using 'name' key instead of 'account' are handled."""
        builder = FinancialStatementBuilder(grouping_with_accounts)
        result = builder.build()

        entry_map = {e.lead_sheet_ref: e for e in result.mapping_trace}
        n_entry = entry_map["N"]
        # Third account in N uses "name" key instead of "account"
        names = [acct.account_name for acct in n_entry.accounts]
        assert "Depreciation Expense" in names


class TestMappingTracePDF:
    """Tests for mapping trace in PDF generation."""

    def test_pdf_with_mapping_trace(self, grouping_with_accounts):
        """PDF generates with populated accounts, bytes > 0."""
        from pdf_generator import generate_financial_statements_pdf

        builder = FinancialStatementBuilder(grouping_with_accounts, entity_name="Trace Corp")
        statements = builder.build()
        assert len(statements.mapping_trace) > 0

        pdf_bytes = generate_financial_statements_pdf(statements)
        assert len(pdf_bytes) > 100
        assert isinstance(pdf_bytes, bytes)


class TestMappingTraceExcel:
    """Tests for mapping trace in Excel generation."""

    def test_excel_has_mapping_trace_tab(self, grouping_with_accounts):
        """Excel has 'Mapping Trace' in sheetnames."""
        from openpyxl import load_workbook

        from excel_generator import generate_financial_statements_excel

        builder = FinancialStatementBuilder(grouping_with_accounts, entity_name="Trace Corp")
        statements = builder.build()

        excel_bytes = generate_financial_statements_excel(statements)
        wb = load_workbook(BytesIO(excel_bytes))
        assert "Mapping Trace" in wb.sheetnames
