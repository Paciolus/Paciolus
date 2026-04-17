"""Tests for loan amortization XLSX and PDF exports (Sprint 672)."""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from loan_amortization_engine import (
    ExtraPayment,
    LoanConfig,
    LoanMethod,
    generate_amortization_schedule,
)
from loan_amortization_excel import build_amortization_workbook
from pdf.sections.loan_amortization import build_amortization_pdf


def _baseline_result(**overrides):
    cfg = LoanConfig(
        principal=Decimal("200000"),
        annual_rate=Decimal("0.06"),
        term_periods=360,
        frequency="monthly",
        start_date=date(2026, 1, 1),
    )
    for k, v in overrides.items():
        setattr(cfg, k, v)
    return generate_amortization_schedule(cfg)


class TestXlsxExport:
    def test_workbook_has_three_sheets(self):
        result = _baseline_result()
        wb_bytes = build_amortization_workbook(result)
        wb = load_workbook(io.BytesIO(wb_bytes))
        assert wb.sheetnames == ["Schedule", "Annual Summary", "Inputs"]

    def test_schedule_row_count_matches_result(self):
        result = _baseline_result()
        wb = load_workbook(io.BytesIO(build_amortization_workbook(result)))
        schedule = wb["Schedule"]
        # header + 360 data rows + blank + totals row
        assert schedule.max_row == 1 + 360 + 2

    def test_frozen_header(self):
        result = _baseline_result()
        wb = load_workbook(io.BytesIO(build_amortization_workbook(result)))
        assert wb["Schedule"].freeze_panes == "A2"

    def test_currency_formatting(self):
        result = _baseline_result()
        wb = load_workbook(io.BytesIO(build_amortization_workbook(result)))
        schedule = wb["Schedule"]
        # Beginning balance column
        assert '"$"#,##0.00' in schedule["C2"].number_format

    def test_total_interest_cell_matches_json(self):
        result = _baseline_result()
        wb = load_workbook(io.BytesIO(build_amortization_workbook(result)))
        schedule = wb["Schedule"]
        # Locate the totals row (first column == "Totals")
        total_row = None
        for row in schedule.iter_rows(min_col=1, max_col=1, values_only=False):
            if row[0].value == "Totals":
                total_row = row[0].row
                break
        assert total_row is not None
        # Interest column (5) must equal the JSON total_interest
        cell_value = schedule.cell(row=total_row, column=5).value
        assert Decimal(str(cell_value)).quantize(Decimal("0.01")) == result.total_interest

    def test_inputs_sheet_echoes_config(self):
        result = _baseline_result()
        wb = load_workbook(io.BytesIO(build_amortization_workbook(result)))
        inputs = wb["Inputs"]
        field_values = {
            inputs.cell(row=r, column=1).value: inputs.cell(row=r, column=2).value
            for r in range(2, inputs.max_row + 1)
            if inputs.cell(row=r, column=1).value
        }
        assert field_values["Principal"] == 200000.0
        assert field_values["Term (periods)"] == 360
        assert field_values["Payment Frequency"] == "monthly"

    def test_extra_payments_render(self):
        cfg = LoanConfig(
            principal=Decimal("50000"),
            annual_rate=Decimal("0.05"),
            term_periods=60,
            frequency="monthly",
            extra_payments=[ExtraPayment(period_number=12, amount=Decimal("2500"))],
        )
        result = generate_amortization_schedule(cfg)
        wb = load_workbook(io.BytesIO(build_amortization_workbook(result)))
        inputs = wb["Inputs"]
        labels = [inputs.cell(row=r, column=1).value for r in range(1, inputs.max_row + 1)]
        assert "Extra Payments" in labels


class TestPdfExport:
    def test_pdf_starts_with_magic(self):
        result = _baseline_result()
        pdf = build_amortization_pdf(result)
        assert pdf[:4] == b"%PDF"

    def test_pdf_paginates_for_large_schedule(self):
        """30-year monthly = 360 rows must flow across multiple PDF pages."""
        result = _baseline_result()
        pdf = build_amortization_pdf(result)
        # ReportLab writes one /Type /Page entry per page — must be >1
        page_count = pdf.count(b"/Type /Page\n")
        assert page_count > 1, f"Expected multi-page output, got {page_count}"

    def test_pdf_large_enough_to_paginate_360_rows(self):
        """30-year monthly schedule must produce a meaningful PDF > 20 KB.

        Not a tight bound — just confirms the document isn't a single-page stub.
        """
        result = _baseline_result()
        pdf = build_amortization_pdf(result)
        assert len(pdf) > 20_000

    def test_pdf_handles_interest_only_method(self):
        cfg = LoanConfig(
            principal=Decimal("100000"),
            annual_rate=Decimal("0.07"),
            term_periods=24,
            frequency="monthly",
            method=LoanMethod.INTEREST_ONLY,
        )
        result = generate_amortization_schedule(cfg)
        pdf = build_amortization_pdf(result)
        assert pdf[:4] == b"%PDF"


class TestRouteRegistration:
    def test_xlsx_and_pdf_routes_registered(self):
        from routes.loan_amortization import router

        paths = {route.path for route in router.routes}
        assert "/audit/loan-amortization/export.xlsx" in paths
        assert "/audit/loan-amortization/export.pdf" in paths
