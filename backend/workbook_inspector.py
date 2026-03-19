"""Fast inspection of Excel workbooks to retrieve sheet metadata without full processing."""

import gc
import io
from dataclasses import asdict, dataclass

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from security_utils import log_secure_operation


@dataclass
class SheetInfo:
    """Metadata for a single sheet in a workbook."""

    name: str
    row_count: int
    column_count: int
    columns: list[str]  # First row headers
    has_data: bool

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class WorkbookInfo:
    """Metadata for an Excel workbook."""

    filename: str
    sheet_count: int
    sheets: list[SheetInfo]
    total_rows: int
    is_multi_sheet: bool
    format: str  # 'xlsx' or 'xls'

    def to_dict(self) -> dict:
        return {
            "filename": self.filename,
            "sheet_count": self.sheet_count,
            "sheets": [s.to_dict() for s in self.sheets],
            "total_rows": self.total_rows,
            "is_multi_sheet": self.is_multi_sheet,
            "format": self.format,
        }


def inspect_workbook(file_bytes: bytes, filename: str = "") -> WorkbookInfo:
    """Quickly inspect an Excel workbook and return sheet metadata."""
    log_secure_operation("inspect_workbook", f"Inspecting workbook: {filename}")

    filename_lower = filename.lower()
    is_xlsx = filename_lower.endswith(".xlsx")
    is_xls = filename_lower.endswith(".xls") and not is_xlsx

    is_ods = filename_lower.endswith(".ods")

    if not (is_xlsx or is_xls or is_ods):
        # Try to detect format from content
        if file_bytes[:4] == b"PK\x03\x04":  # ZIP signature (xlsx or ods)
            from shared.ods_parser import _is_ods_zip

            if _is_ods_zip(file_bytes):
                is_ods = True
            else:
                is_xlsx = True
        elif file_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # OLE signature (xls)
            is_xls = True
        else:
            raise ValueError("File is not a recognized spreadsheet format (.xlsx, .xls, or .ods)")

    buffer = io.BytesIO(file_bytes)
    sheets: list[SheetInfo] = []

    try:
        if is_ods:
            sheets = _inspect_ods(buffer, filename)
            file_format = "ods"
        elif is_xlsx:
            # Use openpyxl for .xlsx (faster metadata access)
            sheets = _inspect_xlsx(buffer, filename)
            file_format = "xlsx"
        else:
            # Use pandas with xlrd for .xls
            sheets = _inspect_xls(buffer, filename)
            file_format = "xls"

    except InvalidFileException as e:
        log_secure_operation("inspect_workbook_error", f"Invalid Excel file: {e}")
        raise ValueError(f"Invalid Excel file: {e}")
    except (KeyError, TypeError, OSError) as e:
        log_secure_operation("inspect_workbook_error", f"Error inspecting workbook: {e}")
        raise ValueError(f"Error inspecting workbook: {e}")
    finally:
        buffer.close()
        del buffer
        gc.collect()

    total_rows = sum(s.row_count for s in sheets)
    is_multi_sheet = len(sheets) > 1

    workbook_info = WorkbookInfo(
        filename=filename,
        sheet_count=len(sheets),
        sheets=sheets,
        total_rows=total_rows,
        is_multi_sheet=is_multi_sheet,
        format=file_format,
    )

    log_secure_operation("inspect_workbook_complete", f"Found {len(sheets)} sheet(s), {total_rows} total rows")

    return workbook_info


def _inspect_xlsx(buffer: io.BytesIO, filename: str) -> list[SheetInfo]:
    """Inspect an .xlsx file using openpyxl in read-only mode."""
    sheets: list[SheetInfo] = []

    # Use read_only mode for faster inspection
    wb = load_workbook(buffer, read_only=True, data_only=True)

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get dimensions
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0

            # Get column headers (first row)
            columns: list[str] = []
            if row_count > 0:
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if first_row:
                    columns = [str(cell) if cell is not None else f"Column {i + 1}" for i, cell in enumerate(first_row)]

            has_data = row_count > 1 or (row_count == 1 and col_count > 0)

            sheets.append(
                SheetInfo(
                    name=sheet_name,
                    row_count=max(0, row_count - 1) if row_count > 0 else 0,  # Exclude header
                    column_count=col_count,
                    columns=columns,
                    has_data=has_data,
                )
            )

            log_secure_operation("inspect_sheet", f"Sheet '{sheet_name}': {row_count} rows, {col_count} cols")
    finally:
        wb.close()
        del wb
        gc.collect()

    return sheets


def _inspect_xls(buffer: io.BytesIO, filename: str) -> list[SheetInfo]:
    """Inspect an .xls file using xlrd sheet metadata (no DataFrame materialization)."""
    sheets: list[SheetInfo] = []

    excel_file = pd.ExcelFile(buffer, engine="xlrd")

    try:
        book = excel_file.book
        for sheet_name in excel_file.sheet_names:
            sheet = book.sheet_by_name(sheet_name)
            total_rows = sheet.nrows
            col_count = sheet.ncols

            # First row is the header; data rows exclude it
            columns: list[str] = []
            if total_rows > 0:
                columns = [
                    str(v) if v not in (None, "") else f"Column {i + 1}" for i, v in enumerate(sheet.row_values(0))
                ]

            row_count = max(0, total_rows - 1) if total_rows > 0 else 0
            has_data = row_count > 0

            sheets.append(
                SheetInfo(
                    name=sheet_name,
                    row_count=row_count,
                    column_count=col_count,
                    columns=columns,
                    has_data=has_data,
                )
            )

            log_secure_operation("inspect_sheet", f"Sheet '{sheet_name}': {row_count} rows, {col_count} cols")

    finally:
        excel_file.close()

    return sheets


def _inspect_ods(buffer: io.BytesIO, filename: str) -> list[SheetInfo]:
    """Inspect an .ods file using odfpy DOM (no DataFrame materialization)."""
    from odf.namespaces import TABLENS
    from odf.opendocument import load as load_ods
    from odf.table import Table, TableCell, TableRow
    from odf.text import P

    sheets: list[SheetInfo] = []

    doc = load_ods(buffer)
    try:
        for tbl in doc.spreadsheet.getElementsByType(Table):
            sheet_name = tbl.getAttribute("name") or tbl.getAttrNS(TABLENS, "name") or "Sheet"

            rows = tbl.getElementsByType(TableRow)

            # Count data rows, respecting table:number-rows-repeated but
            # skipping trailing empty rows (ODS pads sheets to 2^20 rows).
            def _row_has_content(row) -> bool:
                for cell in row.getElementsByType(TableCell):
                    if cell.getElementsByType(P):
                        return True
                return False

            total_rows = 0
            last_content_row = 0
            for idx, row in enumerate(rows):
                repeated = row.getAttrNS(TABLENS, "number-rows-repeated")
                n = int(repeated) if repeated else 1
                total_rows += n
                if _row_has_content(row):
                    last_content_row = total_rows

            # Trim to last row with content
            total_rows = last_content_row

            # Extract header columns from first row
            columns: list[str] = []
            is_placeholder: list[bool] = []
            col_count = 0
            if rows and total_rows > 0:
                first_row = rows[0]
                for cell in first_row.getElementsByType(TableCell):
                    repeated = cell.getAttrNS(TABLENS, "number-columns-repeated")
                    n = int(repeated) if repeated else 1
                    paragraphs = cell.getElementsByType(P)
                    text = "".join(str(p) for p in paragraphs).strip() if paragraphs else ""
                    for _ in range(n):
                        if text:
                            columns.append(text)
                            is_placeholder.append(False)
                        else:
                            columns.append(f"Column {len(columns) + 1}")
                            is_placeholder.append(True)
                # Trim trailing placeholder columns
                while is_placeholder and is_placeholder[-1]:
                    columns.pop()
                    is_placeholder.pop()
                col_count = len(columns)

            row_count = max(0, total_rows - 1) if total_rows > 0 else 0
            has_data = row_count > 0

            sheets.append(
                SheetInfo(
                    name=sheet_name,
                    row_count=row_count,
                    column_count=col_count,
                    columns=columns,
                    has_data=has_data,
                )
            )

            log_secure_operation("inspect_sheet", f"Sheet '{sheet_name}': {row_count} rows, {col_count} cols")

    finally:
        del doc

    return sheets


def is_excel_file(filename: str) -> bool:
    """Check if a filename indicates an Excel or ODS spreadsheet file."""
    lower = filename.lower()
    return lower.endswith(".xlsx") or lower.endswith(".xls") or lower.endswith(".ods")


def is_multi_sheet_file(file_bytes: bytes, filename: str) -> bool:
    """
    Quick check if an Excel file has multiple sheets.
    Returns False for non-Excel files.
    """
    if not is_excel_file(filename):
        return False

    try:
        info = inspect_workbook(file_bytes, filename)
        return info.is_multi_sheet
    except (ValueError, OSError):
        return False
