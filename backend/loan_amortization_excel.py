"""XLSX export for the loan amortization tool (Sprint 672).

Produces a three-sheet workbook:
  - Schedule       Period-by-period amortization with currency formatting
  - Annual Summary Year-level roll-up of payments, interest, principal
  - Inputs         Echoed configuration, total interest, total payments, payoff

All currency columns use ``$#,##0.00`` formatting so totals paste cleanly into
engagement workpapers. The Schedule header row is frozen.
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_generator import ExcelColors
from loan_amortization_engine import AmortizationResult

CURRENCY_FORMAT = '"$"#,##0.00'
PERCENT_FORMAT = "0.0000%"

_HEADER_FILL = PatternFill("solid", fgColor=ExcelColors.OBSIDIAN)
_HEADER_FONT = Font(bold=True, color=ExcelColors.OATMEAL, size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_TOTAL_FONT = Font(bold=True, color=ExcelColors.OBSIDIAN, size=11)


def _to_float(value: Decimal | str | int | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _apply_header(row_cells) -> None:
    for cell in row_cells:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize(ws: Worksheet, min_widths: dict[int, int]) -> None:
    """Pick column widths from max content length, bounded below by `min_widths`."""
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = min_widths.get(col_idx, 10)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 32)


def _build_schedule_sheet(ws: Worksheet, result: AmortizationResult) -> None:
    headers = [
        "Period",
        "Payment Date",
        "Beginning Balance",
        "Scheduled Payment",
        "Interest",
        "Principal",
        "Extra Principal",
        "Ending Balance",
    ]
    ws.append(headers)
    _apply_header(ws[1])
    ws.freeze_panes = "A2"

    for entry in result.schedule:
        ws.append(
            [
                entry.period_number,
                entry.payment_date.isoformat() if entry.payment_date else "",
                _to_float(entry.beginning_balance),
                _to_float(entry.scheduled_payment),
                _to_float(entry.interest),
                _to_float(entry.principal),
                _to_float(entry.extra_principal),
                _to_float(entry.ending_balance),
            ]
        )

    for col in range(3, 9):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = CURRENCY_FORMAT

    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="Totals").font = _TOTAL_FONT
    ws.cell(row=total_row, column=5, value=_to_float(result.total_interest)).number_format = CURRENCY_FORMAT
    ws.cell(row=total_row, column=5).font = _TOTAL_FONT
    ws.cell(row=total_row, column=8, value=_to_float(result.total_payments)).number_format = CURRENCY_FORMAT
    ws.cell(row=total_row, column=8).font = _TOTAL_FONT

    _autosize(
        ws,
        min_widths={1: 7, 2: 12, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 14},
    )


def _build_annual_summary_sheet(ws: Worksheet, result: AmortizationResult) -> None:
    headers = ["Year", "Calendar Year", "Total Payments", "Total Interest", "Total Principal", "Ending Balance"]
    ws.append(headers)
    _apply_header(ws[1])

    for entry in result.annual_summary:
        ws.append(
            [
                entry.year_index,
                entry.calendar_year if entry.calendar_year else "",
                _to_float(entry.total_payments),
                _to_float(entry.total_interest),
                _to_float(entry.total_principal),
                _to_float(entry.ending_balance),
            ]
        )

    for col in range(3, 7):
        letter = get_column_letter(col)
        for row in range(2, ws.max_row + 1):
            ws[f"{letter}{row}"].number_format = CURRENCY_FORMAT

    _autosize(ws, min_widths={1: 6, 2: 14, 3: 16, 4: 16, 5: 16, 6: 16})


def _build_inputs_sheet(ws: Worksheet, result: AmortizationResult) -> None:
    cfg = result.config
    ws.append(["Field", "Value"])
    _apply_header(ws[1])

    rows: list[tuple[str, object]] = [
        ("Principal", _to_float(cfg.principal)),
        ("Annual Rate", _to_float(cfg.annual_rate)),
        ("Term (periods)", cfg.term_periods),
        ("Payment Frequency", cfg.frequency),
        ("Method", cfg.method.value),
        ("Start Date", cfg.start_date.isoformat() if cfg.start_date else ""),
        ("Balloon Amount", _to_float(cfg.balloon_amount) if cfg.balloon_amount else ""),
        ("Total Interest", _to_float(result.total_interest)),
        ("Total Payments", _to_float(result.total_payments)),
        ("Payoff Date", result.payoff_date.isoformat() if result.payoff_date else ""),
    ]
    for label, value in rows:
        ws.append([label, value])

    # Apply formatting per-field
    ws["B2"].number_format = CURRENCY_FORMAT  # Principal
    ws["B3"].number_format = PERCENT_FORMAT  # Annual Rate
    ws["B8"].number_format = CURRENCY_FORMAT  # Balloon Amount
    ws["B9"].number_format = CURRENCY_FORMAT  # Total Interest
    ws["B10"].number_format = CURRENCY_FORMAT  # Total Payments

    if cfg.extra_payments:
        start = ws.max_row + 2
        ws.cell(row=start, column=1, value="Extra Payments").font = _TOTAL_FONT
        ws.cell(row=start + 1, column=1, value="Period")
        ws.cell(row=start + 1, column=2, value="Amount")
        _apply_header([ws.cell(row=start + 1, column=1), ws.cell(row=start + 1, column=2)])
        for i, extra in enumerate(cfg.extra_payments, start=2):
            ws.cell(row=start + i, column=1, value=extra.period_number)
            cell = ws.cell(row=start + i, column=2, value=_to_float(extra.amount))
            cell.number_format = CURRENCY_FORMAT

    if cfg.rate_changes:
        start = ws.max_row + 2
        ws.cell(row=start, column=1, value="Rate Changes").font = _TOTAL_FONT
        ws.cell(row=start + 1, column=1, value="Period")
        ws.cell(row=start + 1, column=2, value="New Annual Rate")
        _apply_header([ws.cell(row=start + 1, column=1), ws.cell(row=start + 1, column=2)])
        for i, chg in enumerate(cfg.rate_changes, start=2):
            ws.cell(row=start + i, column=1, value=chg.period_number)
            cell = ws.cell(row=start + i, column=2, value=_to_float(chg.new_annual_rate))
            cell.number_format = PERCENT_FORMAT

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20


def build_amortization_workbook(result: AmortizationResult) -> bytes:
    """Render the three-sheet XLSX workbook and return bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    schedule_ws = wb.create_sheet("Schedule")
    annual_ws = wb.create_sheet("Annual Summary")
    inputs_ws = wb.create_sheet("Inputs")

    _build_schedule_sheet(schedule_ws, result)
    _build_annual_summary_sheet(annual_ws, result)
    _build_inputs_sheet(inputs_ws, result)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
